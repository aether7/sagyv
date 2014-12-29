# -*- coding: utf-8 -*-
from reportlab.platypus import BaseDocTemplate, PageBreak
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment

class TemplateBasePDF(BaseDocTemplate):

    """
         Parametros Posibles

         filename,
         pagesize=defaultPageSize,
         pageTemplates=[],
         showBoundary=0,
         leftMargin=inch,
         rightMargin=inch,
         topMargin=inch,
         bottomMargin=inch,
         allowSplitting=1,
         title=None,
         author=None,
         _pageBreakQuick=1,
        encrypt=None
    """

    def __init__(self, *args, **kwargs):
        BaseDocTemplate.__init__(self, *args, **kwargs)

        self.estilos = getSampleStyleSheet()
        self.contenido = []

    def saltar_pagina(self):
        self.contenido.append(PageBreak())

    def generar_documento(self):
        self.build(self.contenido)


class ConsumoClientesComerciales(object):

    def __init__(self):
        self.nombre = "consumo_clientes_comerciales"
        self.productos = []
        self.clientes = []
        self.wb = Workbook(encoding="utf-8")
        self.matriz = {}

    def set_productos(self,productos):
        self.productos   = productos

    def set_clientes(self, clientes):
        self.clientes = (c.nombre for c in clientes)

    def construir_reporte(self, datos):
        hoja = self.wb.create_sheet(0, "Consumo")

        cont_col = ord("B")

        for producto in self.productos:
            str_col = str(chr(cont_col)+"1")

            hoja.cell(str_col).value = producto.nombre

            self.matriz[producto.id] = str(chr(cont_col))
            cont_col += 1

        cont_fila = 2

        for nombre in self.clientes:
            str_fila = "A"+str(cont_fila)

            hoja.cell(str_fila).value = nombre
            hoja.cell

            self.matriz[nombre] = str(cont_fila)
            cont_fila += 1

        for consumo in datos:
            columna = self.matriz[consumo.id_producto]
            fila = self.matriz[consumo.nombre]
            hoja.cell(columna+fila).value = str(consumo.suma_monto)

        #Metodo especial hecho para los response django
        return save_virtual_workbook(self.wb)


class KilosDeVentasPorChofer(object):

    def __init__(self):
        self.nombre = "kilos_de_ventas_por_chofer"
        self.wb = Workbook(encoding="utf-8")
        self.matriz = {}

    def construir_reporte(self, datos=[], productos=[], choferes=[]):
        hoja = self.wb.create_sheet(0,"Ventas por Kilo")
        hoja.page_setup.horizontalCentered = True
        hoja.page_setup.verticalCentered = True
        cont_col = ord("B")



        #llenado de productos (desde B1 hasta Bn)
        for producto in productos:
            str_col = str(chr(cont_col)+"1").decode("utf-8")

            alignment = Alignment(wrap_text=True, horizontal= "center")
            hoja.cell(str_col).style = hoja.cell(str_col).style.copy(alignment=alignment)
            hoja.cell(str_col).value = (producto.tipo_producto.nombre+"\n"+producto.nombre).encode("utf-8")

            hoja.column_dimensions[chr(cont_col)].width = len(hoja.cell(str_col).value) - 1

            self.matriz[producto.id] = str(chr(cont_col))
            cont_col += 1

        hoja.row_dimensions[1].height = 35

        cont_fila = 2
        nombre_max = 0

        for chofer in choferes:
            str_fila = "A"+str(cont_fila)

            alignment = Alignment(horizontal= "center")
            hoja.cell(str_fila).style = hoja.cell(str_fila).style.copy(alignment=alignment)
            hoja.cell(str_fila).value = chofer.nombre

            if nombre_max < len(chofer.nombre):
                nombre_max = len(chofer.nombre)

            self.matriz[chofer.nombre] = str(cont_fila)
            cont_fila += 1

        hoja.column_dimensions["A"].width = nombre_max

        for kilos_x_chofer in datos:
            columna = self.matriz[kilos_x_chofer.producto_id]
            fila = self.matriz[kilos_x_chofer.trabajador_nombre]

            celda = columna+fila
            alignment = Alignment(horizontal= "center")
            hoja.cell(celda).style = hoja.cell(celda).style.copy(alignment=alignment)
            hoja.cell(celda).value = str(kilos_x_chofer.suma_kilos)


        return save_virtual_workbook(self.wb)

class DetalleCuotasCreditos(object):

    def __init__(self):
        self.nombre = "detalle_cuotas_creditos_clientes"
        self.wb = Workbook(encoding="utf-8")
        self.matriz = {}
        self.nombre_columnas = ("Cliente","Número Tarjeta", "Tipo Cuotas", "Fecha",
                                "Número Cuotas","Cuotas Pagadas", "Monto Pagado",
                                "Cuotas Impagas", "Monto Impago")

    def construir_reporte(self, creditos=[]):

        hoja = self.wb.create_sheet(0,u"Créditos")

        hoja.page_setup.horizontalCentered = True
        hoja.page_setup.verticalCentered = True
        cont_col = ord("A")

        #llenado de productos (desde B1 hasta Bn)

        for columna in self.nombre_columnas:
            str_col = str(chr(cont_col)+"1").decode("utf-8")
            alignment = Alignment(wrap_text=True, horizontal= "center")
            hoja.cell(str_col).style = hoja.cell(str_col).style.copy(alignment=alignment)
            hoja.cell(str_col).value = columna.decode("utf-8")
            hoja.column_dimensions[chr(cont_col)].width = len(hoja.cell(str_col).value) - 1
            cont_col += 1

        hoja.row_dimensions[1].height = 35

        cont_fila = 2
        nombre_max = 0

        for credito in creditos:

            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=1), credito.nombre_cliente)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=2), credito.numero_tarjeta)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=3), credito.tipo_cuotas)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=4), credito.fecha)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=5), credito.numero_cuotas)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=6), credito.cant_cuotas_pagadas)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=7), credito.cuotas_pagadas)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=8), credito.cant_cuotas_impagas)
            self.setear_alinear_celda(hoja.cell(row=cont_fila, column=9), credito.cuotas_impagas)

            if nombre_max < len(credito.nombre_cliente):
                nombre_max = len(credito.nombre_cliente)

            cont_fila += 1

        hoja.column_dimensions["A"].width = nombre_max
        return save_virtual_workbook(self.wb)

    def setear_alinear_celda(self, celda, valor=0):

        alignment = Alignment(horizontal= "center")
        celda.style = celda.style.copy(alignment=alignment)
        celda.value = valor

if __name__ == "__main__":
    print "oli"