# -*- coding: utf-8 -*-
from reportlab.platypus import BaseDocTemplate, PageBreak
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment

class TemplateBasePDF(BaseDocTemplate):

    """
         Parametros Posibles

         filename,
         pagesize=defaultPageSize,
         pageTemplates=[],
         showBoundary=0,
         leftMargin=inch,
         rightMargin=inch,
         topMargin=inch,
         bottomMargin=inch,
         allowSplitting=1,
         title=None,
         author=None,
         _pageBreakQuick=1,
        encrypt=None
    """

    def __init__(self, *args, **kwargs):
        BaseDocTemplate.__init__(self, *args, **kwargs)

        self.estilos = getSampleStyleSheet()
        self.contenido = []

    def saltar_pagina(self):
        self.contenido.append(PageBreak())

    def generar_documento(self):
        self.build(self.contenido)


class ConsumoClientesComerciales(object):

    def __init__(self):
        self.nombre = "consumo_clientes_comerciales"
        self.productos = []
        self.clientes = []
        self.wb = Workbook(encoding="utf-8")
        self.matriz = {}

    def set_productos(self,productos):
        self.productos   = productos

    def set_clientes(self, clientes):
        self.clientes = (c.nombre for c in clientes)

    def construir_reporte(self, datos):
        hoja = self.wb.create_sheet(0, "Consumo")

        cont_col = ord("B")

        for producto in self.productos:
            str_col = str(chr(cont_col)+"1")

            hoja.cell(str_col).value = producto.nombre

            self.matriz[producto.id] = str(chr(cont_col))
            cont_col += 1

        cont_fila = 2

        for nombre in self.clientes:
            str_fila = "A"+str(cont_fila)

            hoja.cell(str_fila).value = nombre
            hoja.cell

            self.matriz[nombre] = str(cont_fila)
            cont_fila += 1

        for consumo in datos:
            columna = self.matriz[consumo.id_producto]
            fila = self.matriz[consumo.nombre]
            hoja.cell(columna+fila).value = str(consumo.suma_monto)

        #Metodo especial hecho para los response django
        return save_virtual_workbook(self.wb)


class KilosDeVentasPorChofer(object):

    def __init__(self):
        self.nombre = "kilos_de_ventas_por_chofer"
        self.wb = Workbook(encoding="utf-8")
        self.matriz = {}

    def construir_reporte(self, datos=[], productos=[], choferes=[]):
        hoja = self.wb.create_sheet(0,"Ventas por Kilo")
        hoja.page_setup.horizontalCentered = True
        hoja.page_setup.verticalCentered = True
        cont_col = ord("B")



        #llenado de productos (desde B1 hasta Bn)
        for producto in productos:
            str_col = str(chr(cont_col)+"1").encode("utf-8")

            alignment = Alignment(wrap_text=True, horizontal= "center")
            hoja.cell(str_col).style = hoja.cell(str_col).style.copy(alignment=alignment)
            hoja.cell(str_col).value = (producto.tipo_producto.nombre+"\n"+producto.nombre).encode("utf-8")

            hoja.column_dimensions[chr(cont_col)].width = len(hoja.cell(str_col).value) - 1

            self.matriz[producto.id] = str(chr(cont_col))
            cont_col += 1

        hoja.row_dimensions[1].height = 35

        cont_fila = 2
        nombre_max = 0

        for chofer in choferes:
            str_fila = "A"+str(cont_fila)

            alignment = Alignment(horizontal= "center")
            hoja.cell(str_fila).style = hoja.cell(str_fila).style.copy(alignment=alignment)
            hoja.cell(str_fila).value = chofer.nombre

            if nombre_max < len(chofer.nombre):
                nombre_max = len(chofer.nombre)

            self.matriz[chofer.nombre] = str(cont_fila)
            cont_fila += 1

        hoja.column_dimensions["A"].width = nombre_max

        for kilos_x_chofer in datos:
            columna = self.matriz[kilos_x_chofer.producto_id]
            fila = self.matriz[kilos_x_chofer.trabajador_nombre]

            celda = columna+fila
            alignment = Alignment(horizontal= "center")
            hoja.cell(celda).style = hoja.cell(celda).style.copy(alignment=alignment)
            hoja.cell(celda).value = str(kilos_x_chofer.suma_kilos)


        return save_virtual_workbook(self.wb)

if __name__ == "__main__":
    print "oli"